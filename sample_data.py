"""샘플 Excel 파일 생성"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def create_sample_excel(path="sample_report.xlsx"):
    wb = openpyxl.Workbook()

    # 월별 매출 시트
    ws = wb.active
    ws.title = "월별매출"

    headers = ["월", "제품A", "제품B", "제품C", "합계", "전월대비(%)"]
    data = [
        ["1월", 12000000, 8500000, 6200000],
        ["2월", 13500000, 9200000, 5800000],
        ["3월", 15000000, 10100000, 7300000],
        ["4월", 14200000, 9800000, 8100000],
        ["5월", 16800000, 11200000, 8900000],
        ["6월", 18500000, 12500000, 9600000],
    ]

    # 헤더
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2E5090")
        cell.alignment = Alignment(horizontal="center")

    # 데이터
    prev_total = None
    for row_idx, row in enumerate(data, 2):
        month, a, b, c = row
        total = a + b + c
        pct = "" if prev_total is None else round((total - prev_total) / prev_total * 100, 1)
        values = [month, a, b, c, total, pct]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="center")
            if col_idx > 1 and col_idx < 6:
                cell.number_format = "#,##0"
            if col_idx == 5:
                cell.font = Font(bold=True)
                cell.number_format = "#,##0"
        prev_total = total

    # 컬럼 너비
    widths = [8, 14, 14, 14, 16, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 팀별 실적 시트
    ws2 = wb.create_sheet("팀별실적")
    team_headers = ["팀명", "목표", "실적", "달성률(%)"]
    team_data = [
        ["영업1팀", 50000000, 53200000],
        ["영업2팀", 45000000, 41800000],
        ["영업3팀", 40000000, 44100000],
        ["온라인팀", 30000000, 35600000],
    ]

    for col, h in enumerate(team_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E7E34")
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(team_data, 2):
        team, goal, actual = row
        rate = round(actual / goal * 100, 1)
        for col_idx, val in enumerate([team, goal, actual, rate], 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="center")
            if col_idx in (2, 3):
                cell.number_format = "#,##0"

    for i, w in enumerate([12, 14, 14, 12], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    wb.save(path)
    print(f"샘플 Excel 생성: {path}")


if __name__ == "__main__":
    create_sample_excel()
