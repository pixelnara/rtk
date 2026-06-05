"""Excel → PDF 보고서 생성"""
import sys
from datetime import datetime
import openpyxl
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
)
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
import os


# 한글 폰트 등록
def register_fonts():
    font_paths = [
        "/usr/share/fonts/truetype/nanum/NanumGothic.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    ]
    for path in font_paths:
        if os.path.exists(path):
            name = "Korean" if "Nanum" in path else "DejaVu"
            pdfmetrics.registerFont(TTFont(name, path))
            return name
    return "Helvetica"


FONT = register_fonts()

# 색상
COLOR_PRIMARY = colors.HexColor("#2E5090")
COLOR_GREEN = colors.HexColor("#1E7E34")
COLOR_LIGHT = colors.HexColor("#EEF2FA")
COLOR_WHITE = colors.white
COLOR_GRAY = colors.HexColor("#888888")
COLOR_RED = colors.HexColor("#CC0000")


def style(name, **kwargs):
    s = ParagraphStyle(name, fontName=FONT, **kwargs)
    return s


def load_excel(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    result = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = []
        for row in ws.iter_rows(values_only=True):
            if any(v is not None for v in row):
                rows.append(list(row))
        result[sheet] = rows
    return result


def fmt_number(val):
    if val is None or val == "":
        return "-"
    try:
        return f"{int(val):,}"
    except (ValueError, TypeError):
        return str(val)


def fmt_pct(val):
    if val is None or val == "":
        return "-"
    try:
        v = float(val)
        sign = "▲" if v > 0 else ("▼" if v < 0 else "")
        return f"{sign}{abs(v):.1f}%"
    except (ValueError, TypeError):
        return str(val)


def build_monthly_table(rows):
    header = rows[0]
    data_rows = rows[1:]

    table_data = [header]
    for row in data_rows:
        formatted = []
        for i, val in enumerate(row):
            if i == 0:
                formatted.append(str(val) if val else "")
            elif i == 5:
                formatted.append(fmt_pct(val))
            else:
                formatted.append(fmt_number(val))
        table_data.append(formatted)

    col_widths = [20*mm, 30*mm, 30*mm, 30*mm, 32*mm, 30*mm]
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), COLOR_PRIMARY),
        ("TEXTCOLOR", (0, 0), (-1, 0), COLOR_WHITE),
        ("FONTNAME", (0, 0), (-1, -1), FONT),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [COLOR_WHITE, COLOR_LIGHT]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("FONTNAME", (4, 1), (4, -1), FONT),
        ("FONTSIZE", (4, 1), (4, -1), 9),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    return t


def build_team_table(rows):
    header = rows[0]
    data_rows = rows[1:]

    table_data = [header]
    for row in data_rows:
        team, goal, actual, rate = row
        rate_val = float(rate) if rate else 0
        formatted = [
            str(team),
            fmt_number(goal),
            fmt_number(actual),
            f"{rate_val:.1f}%",
        ]
        table_data.append(formatted)

    col_widths = [35*mm, 40*mm, 40*mm, 35*mm]
    t = Table(table_data, colWidths=col_widths, repeatRows=1)

    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), COLOR_GREEN),
        ("TEXTCOLOR", (0, 0), (-1, 0), COLOR_WHITE),
        ("FONTNAME", (0, 0), (-1, -1), FONT),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [COLOR_WHITE, COLOR_LIGHT]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]

    # 달성률 색상
    for i, row in enumerate(data_rows, 1):
        rate = float(row[3]) if row[3] else 0
        color = COLOR_GREEN if rate >= 100 else COLOR_RED
        style_cmds.append(("TEXTCOLOR", (3, i), (3, i), color))

    t.setStyle(TableStyle(style_cmds))
    return t


def build_summary(monthly_rows):
    data = monthly_rows[1:]
    totals = [row[4] for row in data if row[4] is not None]
    if not totals:
        return []
    total_sum = sum(totals)
    avg = total_sum / len(totals)
    best_idx = totals.index(max(totals))
    best_month = data[best_idx][0]

    items = [
        ("상반기 총 매출", f"{int(total_sum):,} 원"),
        ("월 평균 매출", f"{int(avg):,} 원"),
        ("최고 실적 월", f"{best_month}"),
    ]

    summary_data = [[k, v] for k, v in items]
    t = Table(summary_data, colWidths=[50*mm, 100*mm])
    t.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), FONT),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BACKGROUND", (0, 0), (0, -1), COLOR_LIGHT),
        ("FONTNAME", (0, 0), (0, -1), FONT),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
    ]))
    return [t]


def generate_pdf(excel_path, output_path="report.pdf"):
    data = load_excel(excel_path)

    doc = SimpleDocTemplate(
        output_path,
        pagesize=A4,
        leftMargin=20*mm, rightMargin=20*mm,
        topMargin=20*mm, bottomMargin=20*mm,
    )

    title_style = style("Title", fontSize=18, textColor=COLOR_PRIMARY,
                        alignment=TA_LEFT, spaceAfter=2*mm)
    subtitle_style = style("Subtitle", fontSize=10, textColor=COLOR_GRAY,
                           alignment=TA_LEFT, spaceAfter=6*mm)
    section_style = style("Section", fontSize=12, textColor=COLOR_PRIMARY,
                          spaceBefore=6*mm, spaceAfter=3*mm)
    date_style = style("Date", fontSize=9, textColor=COLOR_GRAY,
                       alignment=TA_RIGHT)

    story = []

    # 제목
    story.append(Paragraph("2025년 상반기 매출 보고서", title_style))
    story.append(Paragraph("Monthly Sales Report — H1 2025", subtitle_style))
    story.append(Paragraph(f"작성일: {datetime.today().strftime('%Y년 %m월 %d일')}", date_style))
    story.append(HRFlowable(width="100%", thickness=1.5, color=COLOR_PRIMARY, spaceAfter=6*mm))

    # 요약
    if "월별매출" in data:
        story.append(Paragraph("■ 핵심 요약", section_style))
        story.extend(build_summary(data["월별매출"]))
        story.append(Spacer(1, 5*mm))

    # 월별 매출 표
    if "월별매출" in data:
        story.append(Paragraph("■ 월별 매출 현황", section_style))
        story.append(build_monthly_table(data["월별매출"]))
        story.append(Spacer(1, 5*mm))

    # 팀별 실적 표
    if "팀별실적" in data:
        story.append(Paragraph("■ 팀별 목표 달성 현황", section_style))
        story.append(build_team_table(data["팀별실적"]))

    doc.build(story)
    print(f"PDF 생성 완료: {output_path}")


if __name__ == "__main__":
    excel = sys.argv[1] if len(sys.argv) > 1 else "sample_report.xlsx"
    out = sys.argv[2] if len(sys.argv) > 2 else "report.pdf"
    generate_pdf(excel, out)
